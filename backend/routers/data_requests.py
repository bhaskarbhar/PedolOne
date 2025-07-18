import os
import json
import uuid
from fastapi import APIRouter, HTTPException, Depends, Request, UploadFile, File, Form
from datetime import datetime, timedelta
from fastapi.encoders import jsonable_encoder
from pymongo import MongoClient
from dotenv import load_dotenv
from typing import List, Optional
from pydantic import BaseModel
import pandas as pd
import io
from cryptography.fernet import Fernet
import base64

from models import (
    DataAccessRequest, CreateDataRequest, RespondToRequest,
    InterOrgContract, CreateInterOrgContract
)
from helpers import users_collection, user_pii_collection, policies_collection, logs_collection
from jwt_utils import get_current_user, TokenData
from routers.websocket import send_user_update

load_dotenv()

router = APIRouter(prefix="/data-requests", tags=["Data Access Requests"])

MONGO_URL = os.getenv("MONGO_URL")
client = MongoClient(MONGO_URL)
db = client.get_database("PedolOne")

# Collections
data_requests_collection = db.get_collection("data_requests")
inter_org_contracts_collection = db.get_collection("inter_org_contracts")
organizations_collection = db.get_collection("organizations")

# Create indexes
data_requests_collection.create_index("expires_at", expireAfterSeconds=0)
data_requests_collection.create_index([("target_user_id", 1), ("status", 1)])
data_requests_collection.create_index([("requester_org_id", 1), ("status", 1)])

# Add new models for bulk requests
class BulkDataRequest(BaseModel):
    target_org_id: str
    excel_file_id: str
    purpose: List[str]
    retention_window: str = "30 days"
    request_message: Optional[str] = None

class CreateBulkDataRequest(BaseModel):
    target_org_id: str
    selected_users: List[int]  # List of user IDs
    requested_resources: List[str]
    purpose: List[str]
    retention_window: str = "30 days"
    request_message: Optional[str] = None

class ExcelFileMetadata(BaseModel):
    file_id: str
    original_filename: str
    encrypted_content: str
    access_policy: dict
    uploaded_by: int
    uploaded_at: datetime
    expires_at: datetime

# Create a new collection for Excel files
excel_files_collection = db["excel_files"]

# Generate encryption key (in production, use environment variable)
ENCRYPTION_KEY = Fernet.generate_key()
cipher_suite = Fernet(ENCRYPTION_KEY)

def encrypt_file_content(content: bytes) -> str:
    """Encrypt file content"""
    encrypted_content = cipher_suite.encrypt(content)
    return base64.b64encode(encrypted_content).decode('utf-8')

def decrypt_file_content(encrypted_content: str) -> bytes:
    """Decrypt file content"""
    encrypted_bytes = base64.b64decode(encrypted_content.encode('utf-8'))
    return cipher_suite.decrypt(encrypted_bytes)

def get_organization_by_id(org_id: str):
    """Get organization by ID"""
    return organizations_collection.find_one({"org_id": org_id})

def get_user_by_email(email: str):
    """Get user by email"""
    return users_collection.find_one({"email": email})

@router.post("/send-request")
async def send_data_request(
    request_data: CreateDataRequest,
    current_user: TokenData = Depends(get_current_user),
    http_request: Request = None
):
    """Send a data access request to a user"""
    
    # Verify current user is an organization admin
    user = users_collection.find_one({"userid": current_user.user_id})
    if not user or user.get("user_type") != "organization":
        raise HTTPException(status_code=403, detail="Only organization admins can send data requests")
    
    # Get organization details
    org = get_organization_by_id(user.get("organization_id"))
    if not org:
        raise HTTPException(status_code=404, detail="Organization not found")
    
    # Verify target user exists
    target_user = get_user_by_email(request_data.target_user_email)
    if not target_user:
        raise HTTPException(status_code=404, detail="Target user not found")
    
    # Get target organization info
    target_org_id = None
    target_org_name = "Unknown Organization"
    
    if target_user.get("organization_id"):
        target_org = get_organization_by_id(target_user["organization_id"])
        if target_org:
            target_org_id = target_org["org_id"]
            target_org_name = target_org["org_name"]
    else:
        # If user doesn't have organization_id, try to find organization through policies
        user_policies = list(policies_collection.find({"user_id": target_user["userid"]}))
        if user_policies:
            # Get the most recent policy to determine organization
            latest_policy = max(user_policies, key=lambda x: x.get("created_at", datetime.min))
            if latest_policy.get("target_org_id"):
                target_org = get_organization_by_id(latest_policy["target_org_id"])
                if target_org:
                    target_org_id = target_org["org_id"]
                    target_org_name = target_org["org_name"]
            elif latest_policy.get("shared_with"):
                target_org_name = latest_policy["shared_with"]
    
    # If we have a target_org_name but no target_org_id, try to find the org by name
    if target_org_name != "Unknown Organization" and not target_org_id:
        org_by_name = organizations_collection.find_one({"org_name": target_org_name})
        if org_by_name:
            target_org_id = org_by_name["org_id"]
    
    # Check if there's an active contract between the organizations
    if target_org_id:
        # Get all active contracts between the organizations
        active_contracts = list(inter_org_contracts_collection.find({
            "$or": [
                {"source_org_id": org["org_id"], "target_org_id": target_org_id},
                {"source_org_id": target_org_id, "target_org_id": org["org_id"]}
            ],
            "status": "active",
            "approval_status": "approved"
        }))
        
        if not active_contracts:
            raise HTTPException(
                status_code=400, 
                detail="No active inter-organization contracts found. Please establish a contract before sending data requests."
            )
        
        # Check if any contract allows the requested resources and purposes
        contract_allowed_resources = []
        contract_allowed_purposes = {}
        supporting_contracts = []
        
        for active_contract in active_contracts:
            contract_resources = []
            contract_purposes = {}
            
            # Handle both old and new contract structures
            if active_contract.get("resources_allowed"):
                # New structure with ContractResource objects
                for resource in active_contract.get("resources_allowed", []):
                    if isinstance(resource, dict) and "resource_name" in resource:
                        resource_name = resource["resource_name"]
                        contract_resources.append(resource_name)
                        # Store allowed purposes for this resource
                        if "purpose" in resource:
                            contract_purposes[resource_name] = resource["purpose"]
                    else:
                        # Fallback if resource is just a string
                        contract_resources.append(str(resource))
            elif active_contract.get("allowed_resources"):
                # Old structure with simple list
                allowed_resources = active_contract.get("allowed_resources", [])
                if isinstance(allowed_resources, str):
                    contract_resources = [allowed_resources]
                else:
                    contract_resources = allowed_resources
            
            # Check if this contract supports the requested resources
            unauthorized_resources = [r for r in request_data.requested_resources if r not in contract_resources]
            if not unauthorized_resources:
                # This contract supports all requested resources
                contract_allowed_resources.extend(contract_resources)
                contract_allowed_purposes.update(contract_purposes)
                supporting_contracts.append({
                    "contract_id": active_contract["contract_id"],
                    "contract_name": active_contract.get("contract_name", "Legacy Contract"),
                    "contract_type": active_contract.get("contract_type", "data_sharing")
                })
        
        # Remove duplicates
        contract_allowed_resources = list(set(contract_allowed_resources))
        
        if not supporting_contracts:
            raise HTTPException(
                status_code=400,
                detail=f"No active contracts support the requested resources: {', '.join(request_data.requested_resources)}"
            )
        
        # Check if all requested purposes are allowed for each resource
        unauthorized_purposes = []
        for resource in request_data.requested_resources:
            if resource in contract_allowed_purposes:
                resource_allowed_purposes = contract_allowed_purposes[resource]
                for purpose in request_data.purpose:
                    if purpose not in resource_allowed_purposes:
                        unauthorized_purposes.append(f"{purpose} for {resource}")
            else:
                # If no specific purposes defined for this resource in contract, 
                # assume all purposes are allowed (backward compatibility)
                pass
        
        if unauthorized_purposes:
            raise HTTPException(
                status_code=400,
                detail=f"The following purposes are not allowed by the active contracts: {', '.join(unauthorized_purposes)}"
            )
    else:
        raise HTTPException(
            status_code=400,
            detail="Target user's organization could not be determined. Data requests require an active inter-organization contract."
        )
    
    # Check if user has the requested PII data
    user_pii = user_pii_collection.find_one({"user_id": target_user["userid"]})
    if not user_pii:
        raise HTTPException(status_code=404, detail="User has no PII data")
    
    available_resources = [pii["resource"] for pii in user_pii.get("pii", [])]
    missing_resources = [r for r in request_data.requested_resources if r not in available_resources]
    
    if missing_resources:
        raise HTTPException(
            status_code=400, 
            detail=f"User does not have the following PII types: {', '.join(missing_resources)}"
        )
    
    # Create data request
    request_id = str(uuid.uuid4())
    created_at = datetime.utcnow()
    expires_at = created_at + timedelta(days=7)  # 7 days to respond
    
    data_request = DataAccessRequest(
        request_id=request_id,
        requester_org_id=org["org_id"],
        requester_org_name=org["org_name"],
        target_user_id=target_user["userid"],
        target_user_email=target_user["email"],
        target_org_id=target_org_id,
        target_org_name=target_org_name,
        requested_resources=request_data.requested_resources,
        purpose=request_data.purpose,
        retention_window=request_data.retention_window,
        request_message=request_data.request_message,
        created_at=created_at,
        expires_at=expires_at
    )
    
    # Insert into database - exclude id field to let MongoDB generate new _id
    data_request_data = data_request.model_dump(by_alias=True, exclude={"id"})
    result = data_requests_collection.insert_one(data_request_data)
    
    # Send WebSocket notification to target user
    await send_user_update(
        user_id=str(target_user["userid"]),
        update_type="data_request_received",
        data={"request": data_request.model_dump()}
    )
    
    # Log the request
    client_ip = "unknown"
    if http_request:
        forwarded_for = http_request.headers.get("X-Forwarded-For")
        if forwarded_for:
            client_ip = forwarded_for.split(",")[0].strip()
        elif http_request.headers.get("X-Real-IP"):
            client_ip = http_request.headers.get("X-Real-IP")
        elif http_request.client:
            client_ip = http_request.client.host
    
    log_entry = {
        "user_id": target_user["userid"],
        "fintech_name": org["org_name"],
        "resource_name": ", ".join(request_data.requested_resources),
        "purpose": request_data.purpose,
        "log_type": "data_request_sent",
        "ip_address": client_ip,
        "data_source": "organization",  # Changed from "individual" to "organization"
        "created_at": created_at,
        "request_id": request_id,
        "requester_org_id": org["org_id"],  # Org making the request
        "responder_org_id": target_org_id     # Org of the user whose data is being requested (if any)
    }
    logs_collection.insert_one(log_entry)
    
    return {
        "message": "Data access request sent successfully",
        "request_id": request_id,
        "expires_at": expires_at.isoformat()
    }

@router.post("/upload-excel")
async def upload_excel_file(
    file: UploadFile = File(...),
    target_org_id: str = Form(...),
    purpose: str = Form(...),  # JSON string of purposes
    retention_window: str = Form("30 days"),
    current_user: TokenData = Depends(get_current_user)
):
    """Upload Excel file for bulk data request with encryption and access controls"""
    
    # Verify user is from requesting organization
    user = users_collection.find_one({"userid": current_user.user_id})
    if not user or user.get("user_type") != "organization":
        raise HTTPException(status_code=403, detail="Only organization users can upload files")
    
    # Validate file type
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are allowed")
    
    # Read and validate Excel content
    try:
        content = await file.read()
        df = pd.read_excel(io.BytesIO(content))
        
        # Validate required columns
        required_columns = ['email', 'full_name', 'resource_type', 'purpose']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise HTTPException(status_code=400, detail=f"Missing required columns: {missing_columns}")
        
        # Validate data (max 1000 rows)
        if len(df) > 1000:
            raise HTTPException(status_code=400, detail="Maximum 1000 rows allowed per file")
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {str(e)}")
    
    # Encrypt file content
    encrypted_content = encrypt_file_content(content)
    
    # Create access policy
    access_policy = {
        "view_only": True,
        "no_download": True,
        "no_copy": True,
        "no_edit": True,
        "no_print": True,
        "expires_at": (datetime.utcnow() + timedelta(days=30)).isoformat(),
        "allowed_orgs": [target_org_id],
        "uploaded_by": current_user.user_id
    }
    
    # Generate unique file ID
    file_id = str(uuid.uuid4())
    
    # Store file metadata
    file_metadata = {
        "file_id": file_id,
        "original_filename": file.filename,
        "encrypted_content": encrypted_content,
        "access_policy": access_policy,
        "uploaded_by": current_user.user_id,
        "uploaded_at": datetime.utcnow(),
        "expires_at": datetime.utcnow() + timedelta(days=30),
        "target_org_id": target_org_id,
        "purposes": json.loads(purpose) if isinstance(purpose, str) else purpose,
        "retention_window": retention_window,
        "row_count": len(df)
    }
    
    excel_files_collection.insert_one(file_metadata)
    
    # Log the upload
    client_ip = "unknown"
    if hasattr(Request, 'client') and Request.client:
        client_ip = Request.client.host
    
    log_entry = {
        "user_id": current_user.user_id,
        "fintech_name": user.get("organization_id", "Unknown"),
        "resource_name": "excel_bulk_request",
        "purpose": json.loads(purpose) if isinstance(purpose, str) else purpose,
        "log_type": "bulk_data_request_uploaded",
        "ip_address": client_ip,
        "data_source": "organization",
        "created_at": datetime.utcnow(),
        "file_id": file_id,
        "target_org_id": target_org_id,
        "row_count": len(df)
    }
    logs_collection.insert_one(log_entry)
    
    return {
        "message": "Excel file uploaded successfully",
        "file_id": file_id,
        "row_count": len(df),
        "expires_at": file_metadata["expires_at"].isoformat()
    }

@router.get("/view-excel/{file_id}")
async def view_excel_file(
    file_id: str,
    current_user: TokenData = Depends(get_current_user)
):
    """View Excel file securely (no download, no copy, no edit)"""
    
    # Get file metadata
    file_metadata = excel_files_collection.find_one({"file_id": file_id})
    if not file_metadata:
        raise HTTPException(status_code=404, detail="File not found")
    
    # Check access permissions
    user = users_collection.find_one({"userid": current_user.user_id})
    if not user:
        raise HTTPException(status_code=403, detail="User not found")
    
    # Check if user is from allowed organization
    user_org_id = user.get("organization_id")
    if user_org_id not in file_metadata["access_policy"]["allowed_orgs"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Check if file has expired
    if datetime.utcnow() > file_metadata["expires_at"]:
        raise HTTPException(status_code=400, detail="File has expired")
    
    # Decrypt and read file content
    try:
        decrypted_content = decrypt_file_content(file_metadata["encrypted_content"])
        df = pd.read_excel(io.BytesIO(decrypted_content))
        
        # Convert to HTML for secure viewing (no download capability)
        html_content = df.to_html(
            index=False,
            classes=['table', 'table-striped', 'table-bordered'],
            table_id='secure-excel-table'
        )
        
        # Add CSS to prevent selection and copying
        secure_html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Secure Excel Viewer</title>
            <style>
                body {{ 
                    font-family: Arial, sans-serif; 
                    margin: 20px;
                    user-select: none;
                    -webkit-user-select: none;
                    -moz-user-select: none;
                    -ms-user-select: none;
                }}
                .secure-header {{
                    background: #f8f9fa;
                    padding: 15px;
                    border-radius: 8px;
                    margin-bottom: 20px;
                    border-left: 4px solid #dc3545;
                }}
                .secure-header h2 {{
                    color: #dc3545;
                    margin: 0;
                }}
                .secure-header p {{
                    margin: 5px 0;
                    color: #6c757d;
                }}
                table {{
                    width: 100%;
                    border-collapse: collapse;
                    background: white;
                }}
                th, td {{
                    border: 1px solid #dee2e6;
                    padding: 8px 12px;
                    text-align: left;
                }}
                th {{
                    background-color: #f8f9fa;
                    font-weight: bold;
                }}
                .no-select {{
                    -webkit-touch-callout: none;
                    -webkit-user-select: none;
                    -khtml-user-select: none;
                    -moz-user-select: none;
                    -ms-user-select: none;
                    user-select: none;
                }}
                .no-copy {{
                    -webkit-user-select: none;
                    -moz-user-select: none;
                    -ms-user-select: none;
                    user-select: none;
                }}
            </style>
        </head>
        <body class="no-select">
            <div class="secure-header">
                <h2>🔒 Secure Excel Viewer</h2>
                <p><strong>File:</strong> {file_metadata['original_filename']}</p>
                <p><strong>Created:</strong> {file_metadata['created_at'].strftime('%Y-%m-%d %H:%M:%S')}</p>
                <p><strong>Expires:</strong> {file_metadata['expires_at'].strftime('%Y-%m-%d %H:%M:%S')}</p>
                <p><strong>Records:</strong> {file_metadata.get('record_count', 'Unknown')}</p>
                <p style="color: #dc3545; font-weight: bold;">⚠️ This file is view-only. Download, copy, and editing are disabled.</p>
            </div>
            <div class="no-copy">
                {html_content}
            </div>
            <script>
                // Disable right-click context menu
                document.addEventListener('contextmenu', function(e) {{
                    e.preventDefault();
                    return false;
                }});
                
                // Disable keyboard shortcuts for copy, save, print
                document.addEventListener('keydown', function(e) {{
                    if ((e.ctrlKey || e.metaKey) && (e.key === 'c' || e.key === 's' || e.key === 'p')) {{
                        e.preventDefault();
                        return false;
                    }}
                }});
                
                // Disable drag and drop
                document.addEventListener('dragstart', function(e) {{
                    e.preventDefault();
                    return false;
                }});
                
                // Disable text selection
                document.addEventListener('selectstart', function(e) {{
                    e.preventDefault();
                    return false;
                }});
            </script>
        </body>
        </html>
        """
        
        return StreamingResponse(
            io.StringIO(secure_html),
            media_type="text/html",
            headers={
                "Content-Disposition": "inline",
                "X-Frame-Options": "DENY",
                "X-Content-Type-Options": "nosniff",
                "Cache-Control": "no-store, no-cache, must-revalidate, private"
            }
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error reading file: {str(e)}")

@router.get("/bulk-requests/{org_id}")
async def get_bulk_requests(
    org_id: str,
    current_user: TokenData = Depends(get_current_user)
):
    """Get bulk data requests for an organization"""
    
    # Verify user is from this organization
    user = users_collection.find_one({"userid": current_user.user_id})
    if not user or user.get("organization_id") != org_id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Get files where this org is the target
    bulk_requests = list(excel_files_collection.find({
        "target_org_id": org_id
    }, sort=[("created_at", -1)]))
    
    # Format response
    formatted_requests = []
    for req in bulk_requests:
        formatted_requests.append({
            "file_id": req["file_id"],
            "original_filename": req["original_filename"],
            "created_at": req["created_at"].isoformat(),
            "expires_at": req["expires_at"].isoformat(),
            "record_count": req.get("record_count", 0),
            "purposes": req.get("purposes", []),
            "retention_window": req.get("retention_window", "30 days"),
            "is_expired": datetime.utcnow() > req["expires_at"]
        })
    
    return formatted_requests

@router.get("/received/{user_id}")
async def get_received_requests(user_id: int, current_user: TokenData = Depends(get_current_user)):
    """Get all data access requests received by a user"""
    
    # Verify user can access these requests
    if current_user.user_id != user_id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    requests = list(data_requests_collection.find(
        {"target_user_id": user_id},
        sort=[("created_at", -1)]
    ))
    
    # Convert ObjectId to string
    for req in requests:
        req["_id"] = str(req["_id"])
        req["created_at"] = req["created_at"].isoformat()
        req["expires_at"] = req["expires_at"].isoformat()
        if req.get("responded_at"):
            req["responded_at"] = req["responded_at"].isoformat()
    
    return jsonable_encoder(requests)

@router.get("/sent/{org_id}")
async def get_sent_requests(org_id: str, current_user: TokenData = Depends(get_current_user)):
    """Get all data access requests sent by an organization"""
    
    # Verify user is admin of this organization
    user = users_collection.find_one({"userid": current_user.user_id})
    if not user or user.get("organization_id") != org_id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    requests = list(data_requests_collection.find(
        {"requester_org_id": org_id},
        sort=[("created_at", -1)]
    ))
    
    # Convert ObjectId to string
    for req in requests:
        req["_id"] = str(req["_id"])
        req["created_at"] = req["created_at"].isoformat()
        req["expires_at"] = req["expires_at"].isoformat()
        if req.get("responded_at"):
            req["responded_at"] = req["responded_at"].isoformat()
    
    return jsonable_encoder(requests)

@router.get("/org/{org_id}")
async def get_organization_data_requests(org_id: str):
    """Get all data requests for an organization (both sent and received)"""
    # Get organization details
    org = get_organization_by_id(org_id)
    if not org:
        raise HTTPException(status_code=404, detail="Organization not found")
    
    org_name = org["org_name"]
    
    # Get requests where this org is the requester (sent requests)
    sent_requests = list(data_requests_collection.find({
        "requester_org_id": org_id
    }))
    
    # Get requests where this org is the target (received requests)
    # First try to find by target_org_id
    received_requests = list(data_requests_collection.find({
        "target_org_id": org_id
    }))
    
    # Also check for any requests that might have this org as target_org_name
    # This handles cases where target_org_id is null but target_org_name is set
    name_based_requests = list(data_requests_collection.find({
        "target_org_name": org_name
    }))
    
    # Combine both results, avoiding duplicates
    all_received_requests = received_requests + name_based_requests
    # Remove duplicates based on request_id
    seen_ids = set()
    unique_received_requests = []
    for req in all_received_requests:
        if req["request_id"] not in seen_ids:
            seen_ids.add(req["request_id"])
            unique_received_requests.append(req)
    
    received_requests = unique_received_requests
    
    # Combine and sort by created_at (newest first)
    all_requests = sent_requests + received_requests
    all_requests.sort(key=lambda x: x.get("created_at", datetime.min), reverse=True)
    
    # Group bulk requests together
    bulk_request_groups = {}
    individual_requests = []
    
    for req in all_requests:
        if req.get("is_bulk_request") and req.get("bulk_request_id"):
            # This is part of a bulk request
            bulk_id = req["bulk_request_id"]
            if bulk_id not in bulk_request_groups:
                bulk_request_groups[bulk_id] = {
                    "bulk_request_id": bulk_id,
                    "requests": [],
                    "is_requester": req["requester_org_id"] == org_id,
                    "requester_org_id": req["requester_org_id"],
                    "requester_org_name": req["requester_org_name"],
                    "target_org_id": req.get("target_org_id"),
                    "target_org_name": req.get("target_org_name", "Unknown Organization"),
                    "requested_resources": req["requested_resources"],
                    "purpose": req["purpose"],
                    "retention_window": req["retention_window"],
                    "request_message": req.get("request_message"),
                    "created_at": req["created_at"],
                    "expires_at": req["expires_at"],
                    "bulk_request_size": req.get("bulk_request_size", 0)
                }
            bulk_request_groups[bulk_id]["requests"].append(req)
        else:
            # This is an individual request
            individual_requests.append(req)
    
    # Format response
    formatted_requests = []
    
    # Add individual requests
    for req in individual_requests:
        is_requester = req["requester_org_id"] == org_id
        target_org_name = req.get("target_org_name", "Unknown Organization")
        
        formatted_requests.append({
            "request_id": req["request_id"],
            "requester_org_id": req["requester_org_id"],
            "requester_org_name": req["requester_org_name"],
            "target_user_id": req["target_user_id"],
            "target_user_email": req["target_user_email"],
            "target_org_name": target_org_name,
            "requested_resources": req["requested_resources"],
            "purpose": req["purpose"],
            "retention_window": req["retention_window"],
            "status": req["status"],
            "request_message": req.get("request_message"),
            "response_message": req.get("response_message"),
            "created_at": req["created_at"].isoformat(),
            "expires_at": req["expires_at"].isoformat(),
            "responded_at": req.get("responded_at").isoformat() if req.get("responded_at") else None,
            "responded_by": req.get("responded_by"),
            "is_requester": is_requester,
            "is_bulk_request": False
        })
    
    # Add bulk requests as single entries
    for bulk_id, bulk_group in bulk_request_groups.items():
        # Calculate overall status for the bulk request
        statuses = [req["status"] for req in bulk_group["requests"]]
        if all(status == "approved" for status in statuses):
            overall_status = "approved"
        elif all(status == "rejected" for status in statuses):
            overall_status = "rejected"
        elif any(status == "pending" for status in statuses):
            overall_status = "pending"
        else:
            overall_status = "mixed"
        
        # Get the first request for basic info
        first_request = bulk_group["requests"][0]
        
        formatted_requests.append({
            "request_id": bulk_id,  # Use bulk_request_id as the main ID
            "bulk_request_id": bulk_id,
            "requester_org_id": bulk_group["requester_org_id"],
            "requester_org_name": bulk_group["requester_org_name"],
            "target_user_id": None,  # Bulk requests don't have a single target user
            "target_user_email": f"{len(bulk_group['requests'])} users",  # Show count
            "target_org_name": bulk_group["target_org_name"],
            "requested_resources": bulk_group["requested_resources"],
            "purpose": bulk_group["purpose"],
            "retention_window": bulk_group["retention_window"],
            "status": overall_status,
            "request_message": bulk_group["request_message"],
            "response_message": None,  # Bulk requests don't have a single response
            "created_at": bulk_group["created_at"].isoformat(),
            "expires_at": bulk_group["expires_at"].isoformat(),
            "responded_at": None,
            "responded_by": None,
            "is_requester": bulk_group["is_requester"],
            "is_bulk_request": True,
            "bulk_request_size": bulk_group["bulk_request_size"],
            "excel_file_id": bulk_group["requests"][0].get("excel_file_id"),  # Get excel_file_id from first request
            "individual_requests": [
                {
                    "request_id": req["request_id"],
                    "target_user_email": req["target_user_email"],
                    "target_user_name": req.get("target_user_name", ""),
                    "status": req["status"],
                    "responded_at": req.get("responded_at").isoformat() if req.get("responded_at") else None
                }
                for req in bulk_group["requests"]
            ]
        })
    
    # Sort by created_at (newest first)
    formatted_requests.sort(key=lambda x: x["created_at"], reverse=True)
    
    return formatted_requests

@router.post("/respond")
async def respond_to_request(
    response_data: RespondToRequest,
    current_user: TokenData = Depends(get_current_user),
    http_request: Request = None
):
    """Respond to a data access request (approve/reject)"""
    
    # Get the request
    request = data_requests_collection.find_one({"request_id": response_data.request_id})
    if not request:
        raise HTTPException(status_code=404, detail="Request not found")
    
    # Verify user can respond to this request
    # Allow the target user OR an admin of the target organization to respond
    current_user_doc = users_collection.find_one({"userid": current_user.user_id})
    if not current_user_doc:
        raise HTTPException(status_code=403, detail="User not found")
    
    can_respond = False
    
    # Check if current user is the target user
    if request["target_user_id"] == current_user.user_id:
        can_respond = True
    # Check if current user is an admin of the target organization
    elif (current_user_doc.get("user_type") == "organization" and 
          current_user_doc.get("organization_id") and
          request.get("target_org_id") == current_user_doc.get("organization_id")):
        can_respond = True
    # Fallback: check if target_org_name matches current user's organization name
    elif (current_user_doc.get("user_type") == "organization" and
          current_user_doc.get("organization_id")):
        current_org = get_organization_by_id(current_user_doc["organization_id"])
        if current_org and request.get("target_org_name") == current_org["org_name"]:
            can_respond = True
    
    if not can_respond:
        raise HTTPException(status_code=403, detail="Access denied. Only the target user or organization admin can respond to this request.")
    
    # Check if request is still pending
    if request["status"] != "pending":
        raise HTTPException(status_code=400, detail="Request has already been responded to")
    
    # Check if request has expired
    if datetime.utcnow() > request["expires_at"]:
        raise HTTPException(status_code=400, detail="Request has expired")
    
    # Update request
    update_data = {
        "status": response_data.status,
        "response_message": response_data.response_message,
        "responded_at": datetime.utcnow(),
        "responded_by": current_user.user_id
    }
    
    data_requests_collection.update_one(
        {"request_id": response_data.request_id},
        {"$set": update_data}
    )
    
    # If approved, create policies automatically based on inter-organization contracts
    if response_data.status == "approved":
        from routers.policy import create_policy_internal
        from models import UserInputPII
        
        # Get active contract between organizations
        active_contract = inter_org_contracts_collection.find_one({
            "$or": [
                {"source_org_id": request["requester_org_id"], "target_org_id": request.get("target_org_id")},
                {"source_org_id": request.get("target_org_id"), "target_org_id": request["requester_org_id"]}
            ],
            "status": "active"
        })
        
        # Get user's PII data
        user_pii = user_pii_collection.find_one({"user_id": current_user.user_id})
        if user_pii:
            for resource in request["requested_resources"]:
                # Check if resource is allowed by contract
                contract_allowed_resources = []
                
                # Handle both old and new contract structures
                if active_contract.get("resources_allowed"):
                    # New structure with ContractResource objects
                    for contract_resource in active_contract.get("resources_allowed", []):
                        if isinstance(contract_resource, dict) and "resource_name" in contract_resource:
                            contract_allowed_resources.append(contract_resource["resource_name"])
                        else:
                            # Fallback if resource is just a string
                            contract_allowed_resources.append(str(contract_resource))
                elif active_contract.get("allowed_resources"):
                    # Old structure with simple list
                    allowed_resources = active_contract.get("allowed_resources", [])
                    if isinstance(allowed_resources, str):
                        contract_allowed_resources = [allowed_resources]
                    else:
                        contract_allowed_resources = allowed_resources
                
                if active_contract and resource in contract_allowed_resources:
                    pii_entry = next((pii for pii in user_pii.get("pii", []) if pii["resource"] == resource), None)
                    if pii_entry:
                        try:
                            from helpers import decrypt_pii
                            pii_value = decrypt_pii(pii_entry["original"])
                            
                            # Get client IP
                            client_ip = "unknown"
                            if http_request:
                                forwarded_for = http_request.headers.get("X-Forwarded-For")
                                if forwarded_for:
                                    client_ip = forwarded_for.split(",")[0].strip()
                                elif http_request.headers.get("X-Real-IP"):
                                    client_ip = http_request.headers.get("X-Real-IP")
                                elif http_request.client:
                                    client_ip = http_request.client.host
                            
                            # Create policy with contract information
                            policy_input = UserInputPII(pii_value=pii_value, resource=resource)
                            create_policy_internal(
                                policy_input, 
                                user_id=current_user.user_id,
                                ip_address=client_ip,
                                target_org_id=request["requester_org_id"],
                                contract_id=active_contract["contract_id"]
                            )
                        except Exception as e:
                            print(f"Error creating policy for {resource}: {e}")
                else:
                    print(f"Resource {resource} not allowed by contract or no active contract found")
    
    # Send WebSocket notification to requester organization
    await send_user_update(
        user_id=str(request["requester_org_id"]),
        update_type="data_request_responded",
        data={
            "request_id": response_data.request_id,
            "status": response_data.status,
            "response_message": response_data.response_message
        }
    )
    
    # Log the response
    client_ip = "unknown"
    if http_request:
        forwarded_for = http_request.headers.get("X-Forwarded-For")
        if forwarded_for:
            client_ip = forwarded_for.split(",")[0].strip()
        elif http_request.headers.get("X-Real-IP"):
            client_ip = http_request.headers.get("X-Real-IP")
        elif http_request.client:
            client_ip = http_request.client.host

    log_entry = {
        "user_id": request["target_user_id"],  # Always use the target user's ID
        "fintech_name": request["requester_org_name"],
        "resource_name": ", ".join(request["requested_resources"]),
        "purpose": request["purpose"],
        "log_type": f"data_request_{response_data.status}",  # Changed to "data_request_approved" or "data_request_rejected"
        "ip_address": client_ip,
        "data_source": "organization",  # Changed from "individual" to "organization"
        "created_at": datetime.utcnow(),
        "request_id": response_data.request_id,
        "response_status": response_data.status,
        "requester_org_id": request["requester_org_id"],  # Org who requested
        "responder_org_id": request.get("target_org_id")  # Org who accepted/responded (if any)
    }
    logs_collection.insert_one(log_entry)
    
    return {
        "message": f"Request {response_data.status} successfully",
        "request_id": response_data.request_id,
        "status": response_data.status
    }

@router.get("/stats/{user_id}")
async def get_request_stats(user_id: int, current_user: TokenData = Depends(get_current_user)):
    """Get statistics about data requests for a user"""
    
    if current_user.user_id != user_id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Count requests by status
    stats = {
        "total_received": data_requests_collection.count_documents({"target_user_id": user_id}),
        "pending": data_requests_collection.count_documents({"target_user_id": user_id, "status": "pending"}),
        "approved": data_requests_collection.count_documents({"target_user_id": user_id, "status": "approved"}),
        "rejected": data_requests_collection.count_documents({"target_user_id": user_id, "status": "rejected"}),
        "expired": data_requests_collection.count_documents({"target_user_id": user_id, "status": "expired"})
    }
    
    return stats

@router.get("/available-organizations/{org_id}")
async def get_available_organizations_for_requests(
    org_id: str, 
    current_user: TokenData = Depends(get_current_user)
):
    """Get all organizations with active contracts that can receive data requests from this organization"""
    
    # Verify user is admin of this organization
    user = users_collection.find_one({"userid": current_user.user_id})
    if not user or user.get("organization_id") != org_id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Get organization details
    org = get_organization_by_id(org_id)
    if not org:
        raise HTTPException(status_code=404, detail="Organization not found")
    
    # Find all active contracts where this org is the source
    active_contracts = list(inter_org_contracts_collection.find({
        "source_org_id": org_id,
        "status": "active"
    }))
    
    # Also find contracts where this org is the target (bidirectional contracts)
    target_contracts = list(inter_org_contracts_collection.find({
        "target_org_id": org_id,
        "status": "active"
    }))
    
    # Combine and process contracts
    available_organizations = {}
    
    # Process source contracts (this org can send requests to target orgs)
    for contract in active_contracts:
        target_org_id = contract["target_org_id"]
        target_org_name = contract["target_org_name"]
        
        if target_org_id not in available_organizations:
            available_organizations[target_org_id] = {
                "org_id": target_org_id,
                "org_name": target_org_name,
                "allowed_resources": [],
                "allowed_purposes": {},
                "contract_id": contract["contract_id"]
            }
        
        # Add allowed resources and purposes from this contract
        if contract.get("resources_allowed"):
            for resource in contract.get("resources_allowed", []):
                if isinstance(resource, dict) and "resource_name" in resource:
                    resource_name = resource["resource_name"]
                    if resource_name not in available_organizations[target_org_id]["allowed_resources"]:
                        available_organizations[target_org_id]["allowed_resources"].append(resource_name)
                    
                    # Add purposes for this resource
                    if "purpose" in resource:
                        if resource_name not in available_organizations[target_org_id]["allowed_purposes"]:
                            available_organizations[target_org_id]["allowed_purposes"][resource_name] = []
                        available_organizations[target_org_id]["allowed_purposes"][resource_name].extend(resource["purpose"])
        
        elif contract.get("allowed_resources"):
            allowed_resources = contract.get("allowed_resources", [])
            if isinstance(allowed_resources, str):
                allowed_resources = [allowed_resources]
            
            for resource in allowed_resources:
                if resource not in available_organizations[target_org_id]["allowed_resources"]:
                    available_organizations[target_org_id]["allowed_resources"].append(resource)
    
    # Process target contracts (bidirectional - this org can also send requests to source orgs)
    for contract in target_contracts:
        source_org_id = contract["source_org_id"]
        source_org_name = contract["source_org_name"]
        
        if source_org_id not in available_organizations:
            available_organizations[source_org_id] = {
                "org_id": source_org_id,
                "org_name": source_org_name,
                "allowed_resources": [],
                "allowed_purposes": {},
                "contract_id": contract["contract_id"]
            }
        
        # Add allowed resources and purposes from this contract
        if contract.get("resources_allowed"):
            for resource in contract.get("resources_allowed", []):
                if isinstance(resource, dict) and "resource_name" in resource:
                    resource_name = resource["resource_name"]
                    if resource_name not in available_organizations[source_org_id]["allowed_resources"]:
                        available_organizations[source_org_id]["allowed_resources"].append(resource_name)
                    
                    # Add purposes for this resource
                    if "purpose" in resource:
                        if resource_name not in available_organizations[source_org_id]["allowed_purposes"]:
                            available_organizations[source_org_id]["allowed_purposes"][resource_name] = []
                        available_organizations[source_org_id]["allowed_purposes"][resource_name].extend(resource["purpose"])
        
        elif contract.get("allowed_resources"):
            allowed_resources = contract.get("allowed_resources", [])
            if isinstance(allowed_resources, str):
                allowed_resources = [allowed_resources]
            
            for resource in allowed_resources:
                if resource not in available_organizations[source_org_id]["allowed_resources"]:
                    available_organizations[source_org_id]["allowed_resources"].append(resource)
    
    # Remove duplicates from purposes lists
    for org_data in available_organizations.values():
        for resource_name, purposes in org_data["allowed_purposes"].items():
            org_data["allowed_purposes"][resource_name] = list(set(purposes))
    
    return list(available_organizations.values())

@router.get("/approved-data/{request_id}")
async def get_approved_data_pii(
    request_id: str,
    current_user: TokenData = Depends(get_current_user)
):
    """Get PII data for an approved data request"""
    
    # Get the data request
    request = data_requests_collection.find_one({"request_id": request_id})
    if not request:
        raise HTTPException(status_code=404, detail="Data request not found")
    
    # Check if request is approved
    if request["status"] != "approved":
        raise HTTPException(status_code=400, detail="Data request is not approved")
    
    # Verify user is from the requesting organization
    user = users_collection.find_one({"userid": current_user.user_id})
    if not user:
        raise HTTPException(status_code=403, detail="User not found")
    
    # Check if user is from the requesting organization
    if user.get("organization_id") != request["requester_org_id"]:
        raise HTTPException(status_code=403, detail="Access denied. Only the requesting organization can view this data.")
    
    # Get the target user's PII data
    target_user_id = request["target_user_id"]
    user_pii = user_pii_collection.find_one({"user_id": target_user_id})
    
    if not user_pii:
        raise HTTPException(status_code=404, detail="No PII data found for this user")
    
    # Get user details
    target_user = users_collection.find_one({"userid": target_user_id})
    if not target_user:
        raise HTTPException(status_code=404, detail="Target user not found")
    
    # Filter PII data to only include requested resources that were approved
    approved_pii_data = []
    
    for pii_entry in user_pii.get("pii", []):
        if pii_entry["resource"] in request["requested_resources"]:
            try:
                from helpers import decrypt_pii
                decrypted_value = decrypt_pii(pii_entry["original"])
                
                approved_pii_data.append({
                    "resource": pii_entry["resource"],
                    "value": decrypted_value,
                    "tokenized": pii_entry.get("tokenized", "N/A")
                })
            except Exception as e:
                print(f"Error decrypting PII for {pii_entry['resource']}: {e}")
                # Include tokenized version if decryption fails
                approved_pii_data.append({
                    "resource": pii_entry["resource"],
                    "value": "Decryption failed",
                    "tokenized": pii_entry.get("tokenized", "N/A")
                })
    
    # Get active policies for this user and requesting organization
    from routers.policy import policies_collection
    active_policies = list(policies_collection.find({
        "user_id": target_user_id,
        "target_org_id": request["requester_org_id"],
        "is_revoked": {"$ne": True}
    }))
    
    # Format policies
    formatted_policies = []
    for policy in active_policies:
        formatted_policies.append({
            "resource_name": policy["resource_name"],
            "purpose": policy.get("purpose", []),
            "retention_window": policy.get("retention_window", "30 days"),
            "created_at": policy["created_at"].isoformat(),
            "contract_id": policy.get("contract_id")
        })
    
    return {
        "user_info": {
            "full_name": target_user["full_name"],
            "email": target_user["email"]
        },
        "pii_data": approved_pii_data,
        "active_policies": formatted_policies,
        "request_details": {
            "request_id": request["request_id"],
            "purpose": request["purpose"],
            "requested_resources": request["requested_resources"],
            "approved_at": request.get("responded_at").isoformat() if request.get("responded_at") else None
        }
    }

@router.get("/bulk-approved-data/{org_id}")
async def get_bulk_approved_data(
    org_id: str,
    current_user: TokenData = Depends(get_current_user)
):
    """Get encrypted Excel file with all approved data requests for an organization"""
    
    # Verify user is from the requesting organization
    user = users_collection.find_one({"userid": current_user.user_id})
    if not user or user.get("organization_id") != org_id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Get all approved data requests for this organization
    approved_requests = list(data_requests_collection.find({
        "requester_org_id": org_id,
        "status": "approved",
        "expires_at": {"$gt": datetime.utcnow()}
    }))
    
    if not approved_requests:
        raise HTTPException(status_code=404, detail="No approved data requests found")
    
    # Collect data for Excel
    excel_data = []
    
    for request in approved_requests:
        # Get target user info
        target_user = users_collection.find_one({"userid": request["target_user_id"]})
        if not target_user:
            continue
        
        # Get PII data for each requested resource
        for resource in request["requested_resources"]:
            # First find the user's PII document
            user_pii_doc = user_pii_collection.find_one({
                "user_id": request["target_user_id"]
            })
            
            if user_pii_doc and "pii" in user_pii_doc:
                # Find the specific resource within the PII array
                pii_entry = next((pii for pii in user_pii_doc["pii"] if pii["resource"] == resource), None)
                
                if pii_entry:
                    try:
                        # Decrypt the PII data
                        from helpers import decrypt_pii
                        decrypted_value = decrypt_pii(pii_entry["original"])
                        
                        excel_data.append({
                            "email": target_user.get("email", "N/A"),
                            "full_name": target_user.get("full_name", "N/A"),
                            "resource_type": resource,
                            "purpose": ", ".join(request["purpose"]) if isinstance(request["purpose"], list) else request["purpose"],
                            "value": decrypted_value,
                            "request_id": request["request_id"],
                            "requested_at": request["created_at"].strftime("%Y-%m-%d %H:%M:%S"),
                            "expires_at": request["expires_at"].strftime("%Y-%m-%d %H:%M:%S")
                        })
                    except Exception as e:
                        excel_data.append({
                            "email": target_user.get("email", "N/A"),
                            "full_name": target_user.get("full_name", "N/A"),
                            "resource_type": resource,
                            "purpose": ", ".join(request["purpose"]) if isinstance(request["purpose"], list) else request["purpose"],
                            "value": "Decryption failed",
                            "request_id": request["request_id"],
                            "requested_at": request["created_at"].strftime("%Y-%m-%d %H:%M:%S"),
                            "expires_at": request["expires_at"].strftime("%Y-%m-%d %H:%M:%S")
                        })
                else:
                    print(f"No PII data found for user {request['target_user_id']}, resource {resource}")
                    excel_data.append({
                        "email": target_user.get("email", "N/A"),
                        "full_name": target_user.get("full_name", "N/A"),
                        "resource_type": resource,
                        "purpose": ", ".join(request["purpose"]) if isinstance(request["purpose"], list) else request["purpose"],
                        "value": "No data available",
                        "request_id": request["request_id"],
                        "requested_at": request["created_at"].strftime("%Y-%m-%d %H:%M:%S"),
                        "expires_at": request["expires_at"].strftime("%Y-%m-%d %H:%M:%S")
                    })
            else:
                print(f"No PII document found for user {request['target_user_id']}")
                excel_data.append({
                    "email": target_user.get("email", "N/A"),
                    "full_name": target_user.get("full_name", "N/A"),
                    "resource_type": resource,
                    "purpose": ", ".join(request["purpose"]) if isinstance(request["purpose"], list) else request["purpose"],
                    "value": "No PII document found",
                    "request_id": request["request_id"],
                    "requested_at": request["created_at"].strftime("%Y-%m-%d %H:%M:%S"),
                    "expires_at": request["expires_at"].strftime("%Y-%m-%d %H:%M:%S")
                })
    
    if not excel_data:
        raise HTTPException(status_code=404, detail="No data available for export")
    
    # Create encrypted Excel file with granular controls
    try:
        df = pd.DataFrame(excel_data)
        
        # Create Excel writer with security settings
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Approved Data', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Approved Data']
            
            # Add security headers
            worksheet['A1'] = "🔒 SECURE DATA EXPORT - VIEW ONLY"
            worksheet['A2'] = f"Generated on: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}"
            worksheet['A3'] = f"Organization: {org_id}"
            worksheet['A4'] = f"Total Records: {len(excel_data)}"
            worksheet['A5'] = "⚠️ This file contains sensitive PII data. Handle with extreme care."
            worksheet['A6'] = "🚫 NO COPY, NO EDIT, NO DOWNLOAD - WEB VIEW ONLY"
            
            # Style the header row
            from openpyxl.styles import Font, PatternFill, Alignment, Protection
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
            
            # Apply header styling to data headers (row 8)
            for cell in worksheet[8]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Add security warning styling
            warning_font = Font(bold=True, color="DC2626")
            for row in range(1, 7):
                worksheet[f'A{row}'].font = warning_font
            
            # Protect the worksheet - make it read-only
            worksheet.protection.sheet = True
            worksheet.protection.password = "SECURE123"  # In production, use environment variable
            
            # Protect all cells
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.protection = Protection(locked=True, hidden=False)
        
        output.seek(0)
        
        # Encrypt the Excel file content
        encrypted_content = encrypt_file_content(output.getvalue())
        
        # Store encrypted file metadata
        file_id = str(uuid.uuid4())
        file_metadata = {
            "file_id": file_id,
            "original_filename": f"approved_data_{org_id}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "encrypted_content": encrypted_content,
            "access_policy": {
                "view_only": True,
                "no_download": True,
                "no_copy": True,
                "no_edit": True,
                "no_print": True,
                "web_only": True,
                "expires_at": (datetime.utcnow() + timedelta(days=7)).isoformat(),
                "allowed_orgs": [org_id],
                "created_by": current_user.user_id
            },
            "created_by": current_user.user_id,
            "created_at": datetime.utcnow(),
            "expires_at": datetime.utcnow() + timedelta(days=7),
            "org_id": org_id,
            "record_count": len(excel_data)
        }
        
        excel_files_collection.insert_one(file_metadata)
        
        # Log the export
        client_ip = "unknown"
        log_entry = {
            "user_id": current_user.user_id,
            "fintech_name": org_id,
            "resource_name": "bulk_data_export",
            "purpose": "Data export for approved requests",
            "log_type": "bulk_data_export",
            "ip_address": client_ip,
            "data_source": "organization",
            "created_at": datetime.utcnow(),
            "exported_records": len(excel_data),
            "requester_org_id": org_id,
            "file_id": file_id
        }
        logs_collection.insert_one(log_entry)
        
        # Return the file ID for web viewing instead of direct download
        return {
            "message": "Encrypted Excel file created successfully",
            "file_id": file_id,
            "view_url": f"/data-requests/view-excel/{file_id}",
            "expires_at": file_metadata["expires_at"].isoformat(),
            "record_count": len(excel_data)
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating Excel file: {str(e)}")

@router.post("/create-bulk-request")
async def create_bulk_data_request(
    request_data: CreateBulkDataRequest,
    current_user: TokenData = Depends(get_current_user),
    http_request: Request = None
):
    """Create a bulk data request for multiple users"""
    
    # Debug logging
    print(f"Received bulk request data: {request_data}")
    print(f"Current user: {current_user.user_id}")
    
    # Verify current user is an organization admin
    user = users_collection.find_one({"userid": current_user.user_id})
    if not user or user.get("user_type") != "organization":
        raise HTTPException(status_code=403, detail="Only organization admins can send bulk data requests")
    
    # Get organization details
    org = get_organization_by_id(user.get("organization_id"))
    if not org:
        raise HTTPException(status_code=404, detail="Organization not found")
    
    # Verify target organization exists
    target_org = get_organization_by_id(request_data.target_org_id)
    if not target_org:
        raise HTTPException(status_code=404, detail="Target organization not found")
    
    # Check if there's an active contract between the organizations
    active_contracts = list(inter_org_contracts_collection.find({
        "$or": [
            {"source_org_id": org["org_id"], "target_org_id": request_data.target_org_id},
            {"source_org_id": request_data.target_org_id, "target_org_id": org["org_id"]}
        ],
        "status": "active",
        "approval_status": "approved"
    }))
    
    if not active_contracts:
        raise HTTPException(
            status_code=400, 
            detail="No active inter-organization contracts found. Please establish a contract before sending bulk data requests."
        )
    
    # Verify all target users exist and belong to the target organization
    target_users = []
    for user_id in request_data.selected_users:
        target_user = users_collection.find_one({"userid": user_id})
        if not target_user:
            raise HTTPException(status_code=404, detail=f"Target user with ID {user_id} not found")
        
        # Check if user belongs to target organization through multiple methods
        user_belongs_to_target = False
        
        # Method 1: Direct organization_id field
        if target_user.get("organization_id") == request_data.target_org_id:
            user_belongs_to_target = True
        
        # Method 2: Check if user has policies with the target organization
        if not user_belongs_to_target:
            user_policies = list(policies_collection.find({
                "user_id": user_id,
                "$or": [
                    {"target_org_id": request_data.target_org_id},
                    {"shared_with": target_org["org_name"]}
                ]
            }))
            if user_policies:
                user_belongs_to_target = True
        
        if not user_belongs_to_target:
            raise HTTPException(status_code=400, detail=f"User {user_id} does not belong to target organization")
        
        target_users.append(target_user)
    
    # Create bulk request ID
    bulk_request_id = str(uuid.uuid4())
    
    # Create individual data requests for each user
    created_requests = []
    for target_user in target_users:
        request_id = str(uuid.uuid4())
        
        # Calculate expiration date
        retention_days = int(request_data.retention_window.split()[0])
        expires_at = datetime.utcnow() + timedelta(days=retention_days)
        
        # Create data request
        data_request = {
            "request_id": request_id,
            "bulk_request_id": bulk_request_id,  # Link to bulk request
            "requester_org_id": org["org_id"],
            "requester_org_name": org["org_name"],
            "target_org_id": request_data.target_org_id,
            "target_org_name": target_org["org_name"],
            "target_user_id": target_user["userid"],
            "target_user_email": target_user["email"],
            "target_user_name": target_user["full_name"],
            "requested_resources": request_data.requested_resources,
            "purpose": request_data.purpose,
            "retention_window": request_data.retention_window,
            "request_message": request_data.request_message,
            "status": "pending",
            "created_at": datetime.utcnow(),
            "expires_at": expires_at,
            "is_bulk_request": True,  # Mark as bulk request
            "bulk_request_size": len(request_data.selected_users)
        }
        
        data_requests_collection.insert_one(data_request)
        created_requests.append(data_request)
    
    # Log the bulk request creation
    client_ip = http_request.client.host if http_request else "unknown"
    log_entry = {
        "user_id": current_user.user_id,
        "fintech_name": org["org_name"],
        "resource_name": "bulk_data_request",
        "purpose": ", ".join(request_data.purpose),
        "log_type": "bulk_data_request_created",
        "ip_address": client_ip,
        "data_source": "organization",
        "created_at": datetime.utcnow(),
        "bulk_request_id": bulk_request_id,
        "requester_org_id": org["org_id"],
        "target_org_id": request_data.target_org_id,
        "user_count": len(request_data.selected_users),
        "resources_requested": request_data.requested_resources
    }
    logs_collection.insert_one(log_entry)
    
    return {
        "message": f"Bulk data request created successfully for {len(created_requests)} users",
        "bulk_request_id": bulk_request_id,
        "created_requests": len(created_requests),
        "target_organization": target_org["org_name"],
        "expires_at": expires_at.isoformat()
    }

@router.get("/bulk-request/{bulk_request_id}")
async def get_bulk_request_details(
    bulk_request_id: str,
    current_user: TokenData = Depends(get_current_user)
):
    """Get details of a specific bulk request"""
    
    # Get all requests for this bulk request
    requests = list(data_requests_collection.find({"bulk_request_id": bulk_request_id}))
    
    if not requests:
        raise HTTPException(status_code=404, detail="Bulk request not found")
    
    # Verify user has access to this bulk request
    user = users_collection.find_one({"userid": current_user.user_id})
    if not user:
        raise HTTPException(status_code=403, detail="User not found")
    
    # Check if user is from requesting or target organization
    requester_org_id = requests[0]["requester_org_id"]
    target_org_id = requests[0]["target_org_id"]
    
    if user.get("organization_id") not in [requester_org_id, target_org_id]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Group requests by status
    status_counts = {}
    for request in requests:
        status = request["status"]
        status_counts[status] = status_counts.get(status, 0) + 1
    
    return {
        "bulk_request_id": bulk_request_id,
        "requester_org_id": requester_org_id,
        "requester_org_name": requests[0]["requester_org_name"],
        "target_org_id": target_org_id,
        "target_org_name": requests[0]["target_org_name"],
        "total_requests": len(requests),
        "status_counts": status_counts,
        "requested_resources": requests[0]["requested_resources"],
        "purpose": requests[0]["purpose"],
        "retention_window": requests[0]["retention_window"],
        "request_message": requests[0]["request_message"],
        "created_at": requests[0]["created_at"].isoformat(),
        "expires_at": requests[0]["expires_at"].isoformat(),
        "requests": [
            {
                "request_id": req["request_id"],
                "target_user_email": req["target_user_email"],
                "target_user_name": req["target_user_name"],
                "status": req["status"],
                "responded_at": req.get("responded_at").isoformat() if req.get("responded_at") else None
            }
            for req in requests
        ]
    }

@router.post("/approve-bulk-request/{bulk_request_id}")
async def approve_bulk_request(
    bulk_request_id: str,
    current_user: TokenData = Depends(get_current_user),
    http_request: Request = None
):
    """Approve all requests in a bulk request and generate encrypted Excel"""
    
    # Get all requests for this bulk request
    requests = list(data_requests_collection.find({"bulk_request_id": bulk_request_id}))
    
    if not requests:
        raise HTTPException(status_code=404, detail="Bulk request not found")
    
    # Verify user is from target organization
    user = users_collection.find_one({"userid": current_user.user_id})
    if not user or user.get("organization_id") != requests[0]["target_org_id"]:
        raise HTTPException(status_code=403, detail="Only target organization can approve bulk requests")
    
    # Check if all requests are pending
    pending_requests = [req for req in requests if req["status"] == "pending"]
    if not pending_requests:
        raise HTTPException(status_code=400, detail="No pending requests to approve")
    
    # Approve all pending requests
    approved_count = 0
    for request in pending_requests:
        data_requests_collection.update_one(
            {"request_id": request["request_id"]},
            {
                "$set": {
                    "status": "approved",
                    "responded_at": datetime.utcnow(),
                    "responded_by": current_user.user_id
                }
            }
        )
        approved_count += 1
    
    # Generate encrypted Excel file with all approved data
    try:
        # Collect data for Excel
        excel_data = []
        
        print(f"Processing {len(requests)} requests for bulk request {bulk_request_id}")
        
        for request in requests:
            if request["status"] == "approved":
                print(f"Processing approved request {request['request_id']} for user {request['target_user_id']}")
                # Get target user info
                target_user = users_collection.find_one({"userid": request["target_user_id"]})
                if not target_user:
                    print(f"Target user {request['target_user_id']} not found")
                    continue
                
                print(f"Target user found: {target_user.get('email', 'N/A')}")
                print(f"Requested resources: {request['requested_resources']}")
                
                # Get PII data for each requested resource
                for resource in request["requested_resources"]:
                    # First find the user's PII document
                    user_pii_doc = user_pii_collection.find_one({
                        "user_id": request["target_user_id"]
                    })
                    
                    if user_pii_doc and "pii" in user_pii_doc:
                        print(f"Found PII document for user {request['target_user_id']} with {len(user_pii_doc['pii'])} PII entries")
                        print(f"Available resources: {[pii.get('resource', 'N/A') for pii in user_pii_doc['pii']]}")
                        
                        # Find the specific resource within the PII array
                        pii_entry = next((pii for pii in user_pii_doc["pii"] if pii["resource"] == resource), None)
                        
                        if pii_entry:
                            print(f"Found PII entry for resource {resource}")
                            try:
                                # Decrypt the PII data
                                from helpers import decrypt_pii
                                decrypted_value = decrypt_pii(pii_entry["original"])
                                
                                excel_data.append({
                                    "email": target_user.get("email", "N/A"),
                                    "full_name": target_user.get("full_name", "N/A"),
                                    "resource_type": resource,
                                    "purpose": ", ".join(request["purpose"]) if isinstance(request["purpose"], list) else request["purpose"],
                                    "value": decrypted_value,
                                    "request_id": request["request_id"],
                                    "requested_at": request["created_at"].strftime("%Y-%m-%d %H:%M:%S"),
                                    "expires_at": request["expires_at"].strftime("%Y-%m-%d %H:%M:%S")
                                })
                            except Exception as e:
                                print(f"Error decrypting PII for user {request['target_user_id']}, resource {resource}: {e}")
                                excel_data.append({
                                    "email": target_user.get("email", "N/A"),
                                    "full_name": target_user.get("full_name", "N/A"),
                                    "resource_type": resource,
                                    "purpose": ", ".join(request["purpose"]) if isinstance(request["purpose"], list) else request["purpose"],
                                    "value": "Decryption failed",
                                    "request_id": request["request_id"],
                                    "requested_at": request["created_at"].strftime("%Y-%m-%d %H:%M:%S"),
                                    "expires_at": request["expires_at"].strftime("%Y-%m-%d %H:%M:%S")
                                })
                        else:
                            print(f"No PII data found for user {request['target_user_id']}, resource {resource}")
                            excel_data.append({
                                "email": target_user.get("email", "N/A"),
                                "full_name": target_user.get("full_name", "N/A"),
                                "resource_type": resource,
                                "purpose": ", ".join(request["purpose"]) if isinstance(request["purpose"], list) else request["purpose"],
                                "value": "No data available",
                                "request_id": request["request_id"],
                                "requested_at": request["created_at"].strftime("%Y-%m-%d %H:%M:%S"),
                                "expires_at": request["expires_at"].strftime("%Y-%m-%d %H:%M:%S")
                            })
                    else:
                        print(f"No PII document found for user {request['target_user_id']}")
                        excel_data.append({
                            "email": target_user.get("email", "N/A"),
                            "full_name": target_user.get("full_name", "N/A"),
                            "resource_type": resource,
                            "purpose": ", ".join(request["purpose"]) if isinstance(request["purpose"], list) else request["purpose"],
                            "value": "No PII document found",
                            "request_id": request["request_id"],
                            "requested_at": request["created_at"].strftime("%Y-%m-%d %H:%M:%S"),
                            "expires_at": request["expires_at"].strftime("%Y-%m-%d %H:%M:%S")
                        })
        
        if not excel_data:
            print(f"No Excel data collected for bulk request {bulk_request_id}")
            print(f"Total requests: {len(requests)}")
            print(f"Approved requests: {len([r for r in requests if r['status'] == 'approved'])}")
            # Instead of raising an error, create an Excel with available data or placeholder
            excel_data.append({
                "email": "No data available",
                "full_name": "No data available", 
                "resource_type": "No data available",
                "purpose": "No data available",
                "value": "No PII data found for the requested resources",
                "request_id": "N/A",
                "requested_at": datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
                "expires_at": datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
            })
        
        print(f"Collected {len(excel_data)} records for Excel export")
        
        # Create encrypted Excel file
        df = pd.DataFrame(excel_data)
        
        # Create Excel writer with security settings
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Bulk Data Export', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Bulk Data Export']
            
            # Add security headers
            worksheet['A1'] = "🔒 BULK DATA EXPORT - VIEW ONLY"
            worksheet['A2'] = f"Generated on: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}"
            worksheet['A3'] = f"Bulk Request ID: {bulk_request_id}"
            worksheet['A4'] = f"Requester: {requests[0]['requester_org_name']}"
            worksheet['A5'] = f"Target: {requests[0]['target_org_name']}"
            worksheet['A6'] = f"Total Records: {len(excel_data)}"
            worksheet['A7'] = "⚠️ This file contains sensitive PII data. Handle with extreme care."
            worksheet['A8'] = "🚫 NO COPY, NO EDIT, NO DOWNLOAD - WEB VIEW ONLY"
            
            # Style the header row
            from openpyxl.styles import Font, PatternFill, Alignment, Protection
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
            
            # Apply header styling to data headers (row 10)
            for cell in worksheet[10]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Add security warning styling
            warning_font = Font(bold=True, color="DC2626")
            for row in range(1, 9):
                worksheet[f'A{row}'].font = warning_font
            
            # Protect the worksheet - make it read-only
            worksheet.protection.sheet = True
            worksheet.protection.password = "SECURE123"
            
            # Protect all cells
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.protection = Protection(locked=True, hidden=False)
        
        output.seek(0)
        
        # Encrypt the Excel file content
        encrypted_content = encrypt_file_content(output.getvalue())
        
        # Store encrypted file metadata
        file_id = str(uuid.uuid4())
        file_metadata = {
            "file_id": file_id,
            "bulk_request_id": bulk_request_id,
            "original_filename": f"bulk_data_export_{bulk_request_id}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "encrypted_content": encrypted_content,
            "access_policy": {
                "view_only": True,
                "no_download": True,
                "no_copy": True,
                "no_edit": True,
                "no_print": True,
                "web_only": True,
                "expires_at": (datetime.utcnow() + timedelta(days=7)).isoformat(),
                "allowed_orgs": [requests[0]["requester_org_id"]],
                "created_by": current_user.user_id
            },
            "created_by": current_user.user_id,
            "created_at": datetime.utcnow(),
            "expires_at": datetime.utcnow() + timedelta(days=7),
            "org_id": requests[0]["requester_org_id"],
            "record_count": len(excel_data),
            "bulk_request_id": bulk_request_id
        }
        
        excel_files_collection.insert_one(file_metadata)
        
        # Update bulk request with file ID
        data_requests_collection.update_many(
            {"bulk_request_id": bulk_request_id},
            {"$set": {"excel_file_id": file_id}}
        )
        
        # Log the approval and export
        client_ip = http_request.client.host if http_request else "unknown"
        log_entry = {
            "user_id": current_user.user_id,
            "fintech_name": requests[0]["target_org_name"],
            "resource_name": "bulk_data_approval",
            "purpose": "Bulk data request approval and export",
            "log_type": "bulk_data_approved",
            "ip_address": client_ip,
            "data_source": "organization",
            "created_at": datetime.utcnow(),
            "bulk_request_id": bulk_request_id,
            "requester_org_id": requests[0]["requester_org_id"],
            "target_org_id": requests[0]["target_org_id"],
            "approved_requests": approved_count,
            "exported_records": len(excel_data),
            "file_id": file_id
        }
        logs_collection.insert_one(log_entry)
        
        return {
            "message": f"Bulk request approved successfully. {approved_count} requests approved.",
            "bulk_request_id": bulk_request_id,
            "approved_requests": approved_count,
            "excel_file_id": file_id,
            "view_url": f"/data-requests/view-excel/{file_id}",
            "record_count": len(excel_data),
            "expires_at": file_metadata["expires_at"].isoformat()
        }
        
    except Exception as e:
        print(f"Error generating Excel file: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error generating Excel file: {str(e)}")

 